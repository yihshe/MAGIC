import os
import glob
import rasterio
import pandas as pd
import numpy as np
from datetime import datetime
from rasterio.warp import calculate_default_transform, reproject, Resampling
from rasterio.transform import rowcol
from openpyxl import load_workbook

# -------------------------------------------------------------------
# 1. USER SETTINGS
# -------------------------------------------------------------------
# Path to the Excel file with in-situ ESU measurements
EXCEL_FILE = "FRM_Veg_Wytham_20180703_V2.xlsx"
IN_SITU_SHEET = "GroundData"

# Column names in the Excel file
# (For Wytham these columns are in lat/lon even if labeled "Northing Coord." / "Easting Coord.")
LAT_COL = "Northing Coord."   # should be ~51.77
LON_COL = "Easting Coord."    # should be ~-1.33
DATE_COL = "Date (dd/mm/yyyy)" # measurement date (e.g. "3/7/18")
LAI_COL = "LAI"               # Total (true) LAI; typically the sum of LAI_down and LAI_up
LCC_COL = "LCC (g m-2)"
CCC_COL = "CCC (g m-2)"

# Sentinel-2 SAFE folder for the available date:
S2_DATE_STR = "2018-06-29"    # acquisition date
S2_SAFE = "S2A_MSIL2A_20180629T112111_N0500_R037_T30UXC_20230828T055820.SAFE"

# List of bands to extract.
# The tuple for each band contains a file-matching pattern and its native resolution.
S2_BANDS = {
    "B02_BLUE":  ("*B02_10m*.jp2", 10),
    "B03_GREEN": ("*B03_10m*.jp2", 10),
    "B04_RED":   ("*B04_10m*.jp2", 10),
    "B05_RE1":   ("*B05_20m*.jp2", 20),
    "B06_RE2":   ("*B06_20m*.jp2", 20),
    "B07_RE3":   ("*B07_20m*.jp2", 20),
    "B08_NIR1":  ("*B08_10m*.jp2", 10),
    "B8A_NIR2":  ("*B8A_20m*.jp2", 20),
    "B09_WV":    ("*B09_60m*.jp2", 60),
    "B11_SWI1":  ("*B11_20m*.jp2", 20),
    "B12_SWI2":  ("*B12_20m*.jp2", 20),
}

# We choose to downsample everything to 20 m.
TARGET_RES = 20

# -------------------------------------------------------------------
# 2. HELPER FUNCTIONS
# -------------------------------------------------------------------
def find_band_files(safe_folder, band_specs):
    """
    Search the SAFE folder for each band file and return a dictionary:
    band_name -> (full_path_to_jp2, native_resolution)
    """
    band_files = {}
    for b_name, (pattern, native_res) in band_specs.items():
        # Search in the GRANULE/*/IMG_DATA directory
        search_path = os.path.join(safe_folder, "GRANULE", "*", "IMG_DATA", pattern)
        matches = glob.glob(search_path)
        if not matches:
            raise FileNotFoundError(f"No match for pattern {pattern} in {search_path}")
        # If multiple files match, take the first one
        band_files[b_name] = (matches[0], native_res)
    return band_files

def resample_band(band_path, target_res, res_method=Resampling.nearest):
    """
    Open the band file using rasterio, resample to target_res (in meters),
    and return (array, transform, crs) for the resampled data.
    """
    with rasterio.open(band_path) as src:
        dst_transform, width, height = calculate_default_transform(
            src.crs, src.crs, src.width, src.height, *src.bounds, resolution=target_res
        )
        dst_shape = (height, width)
        data_src = src.read(1)
        data_dst = np.zeros(dst_shape, dtype=data_src.dtype)
        reproject(
            source=data_src,
            destination=data_dst,
            src_transform=src.transform,
            src_crs=src.crs,
            dst_transform=dst_transform,
            dst_crs=src.crs,
            resampling=res_method
        )
        return data_dst, dst_transform, src.crs

def get_pixel_value(lat, lon, array_2d, transform, crs, src_crs={"init": "EPSG:4326"}):
    """
    Convert the lat,lon (in EPSG:4326) to the coordinate system of the array,
    then compute the row and column and return the pixel value.
    """
    from rasterio.warp import transform as rio_transform
    x, y = rio_transform(src_crs, crs, [lon], [lat])
    x, y = x[0], y[0]
    r, c = rowcol(transform, x, y, op=round)
    # Check bounds
    if 0 <= r < array_2d.shape[0] and 0 <= c < array_2d.shape[1]:
        return array_2d[r, c]
    else:
        return np.nan

# -------------------------------------------------------------------
# 3. LOAD / FILTER IN-SITU DATA
# -------------------------------------------------------------------
wb = load_workbook(EXCEL_FILE, read_only=True)
if IN_SITU_SHEET not in wb.sheetnames:
    raise ValueError(f"Sheet {IN_SITU_SHEET} not found in {EXCEL_FILE}")
df = pd.read_excel(EXCEL_FILE, sheet_name=IN_SITU_SHEET)

# Drop rows without coordinates; optionally, filter by land cover if desired.
df = df.dropna(subset=[LAT_COL, LON_COL])
# If you want only forest ESUs, you can filter on the "Land Cover" column:
# df = df[df["Land Cover"].str.lower().str.contains("forest")]

# Convert the date column to datetime.
df[DATE_COL] = pd.to_datetime(df[DATE_COL], format="%d/%m/%y", errors="coerce")

# -------------------------------------------------------------------
# 4. PREPARE SENTINEL-2 DATA (Downsampled to 20 m)
# -------------------------------------------------------------------
# 4.1. Find band files in the SAFE folder
band_files = find_band_files(S2_SAFE, S2_BANDS)

# 4.2. Resample each band to 20 m resolution.
s2_date = datetime.strptime(S2_DATE_STR, "%Y-%m-%d")

arrays = {}  # band_name -> (array, transform, crs)

for b_name, (jp2, native_res) in band_files.items():
    arr, tfm, crs = resample_band(jp2, TARGET_RES)
    arrays[b_name] = (arr, tfm, crs)

# -------------------------------------------------------------------
# 5. EXTRACT SPECTRAL VALUES FOR EACH ESU
# -------------------------------------------------------------------
# For each ESU, we sample the pixel at the ESU center.
for b_name in S2_BANDS.keys():
    df[b_name] = np.nan  # prepare a column for each band

for idx, row in df.iterrows():
    lat = float(row[LAT_COL])
    lon = float(row[LON_COL])
    T_meas = row[DATE_COL]
    
    # If measurement date is not valid, skip
    if pd.isnull(T_meas):
        continue
    
    # For each band, extract the pixel value.
    for b_name in S2_BANDS.keys():
        arr, tfm, crs = arrays[b_name]
        df.at[idx, b_name] = get_pixel_value(lat, lon, arr, tfm, crs)

# -------------------------------------------------------------------
# 6. SAVE THE RESULT TO CSV
# -------------------------------------------------------------------
OUT_CSV = "wytham_in_situ_S2_20m_closestdate.csv"
df.to_csv(OUT_CSV, index=False)
print(f"Saved combined ESU and Sentinel-2 data to {OUT_CSV}")
